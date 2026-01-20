"""Pytest fixtures for testing."""

import logging
from io import BytesIO
from unittest.mock import Mock

import pandas as pd
import pytest

logging.basicConfig(level=logging.INFO)


@pytest.fixture
def sample_excel_file():
    """Create a sample Excel file for testing.

    Returns:
        BytesIO: Excel file with sample data.
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["Product", "Quantity", "Price", "Category"])
    ws.append(["Widget A", 10, 50.0, "Electronics"])
    ws.append(["Widget B", 5, 75.0, "Electronics"])
    ws.append(["Widget C", 15, 25.0, "Home"])
    ws.append(["Widget D", 8, 30.0, "Home"])

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    file_stream.name = "test.xlsx"
    return file_stream


@pytest.fixture
def sample_excel_file_multi():
    """Create another sample Excel file for testing multiple file loading.

    Returns:
        BytesIO: Excel file with sample data.
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["Product", "Quantity", "Price", "Category"])
    ws.append(["Widget E", 20, 100.0, "Office"])
    ws.append(["Widget F", 12, 45.0, "Office"])
    ws.append(["Widget G", 7, 60.0, "Garden"])
    ws.append(["Widget H", 3, 90.0, "Garden"])

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    file_stream.name = "test2.xlsx"
    return file_stream


@pytest.fixture
def sample_excel_file_numeric_columns():
    """Create an Excel file with mixed string and numeric column names.

    This simulates real-world files like date-based columns (e.g., 20250131).

    Returns:
        BytesIO: Excel file with numeric column names.
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    # First row with mixed column types - string and integers
    ws.append(["省份", 20250131, 20250228, 20250331])
    ws.append(["北京", 100, 200, 300])
    ws.append(["上海", 150, 250, 350])
    ws.append(["广州", 120, 220, 320])

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    file_stream.name = "numeric_columns.xlsx"
    return file_stream


@pytest.fixture
def sample_dataframe():
    """Create a sample DataFrame for testing.

    Returns:
        pd.DataFrame: Sample DataFrame.
    """
    return pd.DataFrame(
        {
            "Name": ["Alice", "Bob", "Charlie"],
            "Sales": [1000, 1500, 1200],
            "Region": ["North", "South", "East"],
        }
    )


@pytest.fixture
def mock_config():
    """Create a mock configuration.

    Returns:
        Mock: Mock configuration object.
    """
    config = Mock()
    config.deepseek_api_key = "test-api-key"
    config.deepseek_model = "deepseek-chat"
    config.deepseek_base_url = "https://api.deepseek.com"
    config.app_title = "Test App"
    config.app_log_level = "INFO"
    return config


@pytest.fixture
def mock_llm_client():
    """Create a mock LLM client.

    Returns:
        Mock: Mock LLM client.
    """
    client = Mock()
    client.api_key = "test-key"
    client.model = "deepseek-chat"
    client.base_url = "https://api.deepseek.com"
    return client


@pytest.fixture
def sample_csv_file():
    """Create a sample CSV file for testing.

    Returns:
        BytesIO: CSV file with sample data.
    """
    csv_content = """Product,Quantity,Price,Category
Widget A,10,50.0,Electronics
Widget B,5,75.0,Electronics
Widget C,15,25.0,Home
Widget D,8,30.0,Home"""

    file_stream = BytesIO(csv_content.encode("utf-8"))
    file_stream.name = "test.csv"
    return file_stream


@pytest.fixture
def sample_csv_file_multi():
    """Create another sample CSV file for testing multiple file loading.

    Returns:
        BytesIO: CSV file with sample data.
    """
    csv_content = """Product,Quantity,Price,Category
Widget E,20,100.0,Office
Widget F,12,45.0,Office
Widget G,7,60.0,Garden
Widget H,3,90.0,Garden"""

    file_stream = BytesIO(csv_content.encode("utf-8"))
    file_stream.name = "test2.csv"
    return file_stream


@pytest.fixture
def empty_csv_file():
    """Create an empty CSV file for testing.

    Returns:
        BytesIO: Empty CSV file with only headers.
    """
    csv_content = """Product,Quantity,Price,Category"""

    file_stream = BytesIO(csv_content.encode("utf-8"))
    file_stream.name = "empty.csv"
    return file_stream
